import sys
import os
from unittest.mock import MagicMock

sys.path.append(os.getcwd())

from app.services.export import export_price_modifications_excel
from app.models.modification_action import ModificationAction
from app.models.product_master import ProductMaster

def test_export():
    db = MagicMock()
    
    # Mock data
    m1 = MagicMock(spec=ModificationAction)
    m1.action_type = "PRICE_INCREASE"
    m1.old_price = 100.0
    m1.new_price = 120.0
    m1.product = None
    m1.cpl_item = None

    m2 = MagicMock(spec=ModificationAction)
    m2.action_type = "PRICE_INCREASE"
    m2.old_price = 200.0
    m2.new_price = 250.0
    m2.product = None
    m2.cpl_item = None
    
    # Simple query mock
    query = MagicMock()
    query.yield_per.return_value = [m1, m2]
    db.query.return_value.options.return_value.filter.return_value = query
    query.filter.return_value = query
    
    print("Testing Price Change Formulas...")
    wb = export_price_modifications_excel(db, client_id=1, job_id=1, selected_types=["PRICE_INCREASE"])
    
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    from openpyxl import load_workbook
    wb_loaded = load_workbook(tmp_path)
    ws = wb_loaded["Price Increase"]

    print(f"Row 3, Col 14 (Formula): {ws.cell(row=3, column=14).value}")
    assert ws.cell(row=3, column=14).value == "=IFERROR((M3-L3)/L3, 0)"
    
    print(f"Row 4, Col 14 (Formula): {ws.cell(row=4, column=14).value}")
    assert ws.cell(row=4, column=14).value == "=IFERROR((M4-L4)/L4, 0)"

    os.remove(tmp_path)
    print("Verification of formulas successful!")

if __name__ == "__main__":
    test_export()
