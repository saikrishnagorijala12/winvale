
import pytest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook
from tests.conftest import FAKE_ADMIN_USER
from io import BytesIO



class TestExportProducts:

    def _mock_workbook(self):
        """Create a minimal real Workbook for the route to .save()."""
        wb = Workbook()
        ws = wb.active
        ws.append(["col1", "col2"])
        ws.append(["val1", "val2"])
        return wb

    def test_export_all_returns_excel(self, client, mock_db):
        with patch(
            "app.routes.export.export_products_excel",
            return_value=self._mock_workbook(),
        ) as mock_export:
            with patch(
                "app.routes.export.get_master_filename",
                return_value="all_products_2025-01-01.xlsx",
            ) as mock_file:
                resp = client.get("/export/")
        assert resp.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in resp.headers["content-type"]
        )
        mock_export.assert_called_once_with(mock_db, None)
        mock_file.assert_called_once_with(mock_db, None)

    def test_export_by_client_id(self, client, mock_db):
        with patch(
            "app.routes.export.export_products_excel",
            return_value=self._mock_workbook(),
        ) as mock_export:
            with patch(
                "app.routes.export.get_master_filename",
                return_value="client_5_products.xlsx",
            ) as mock_file:
                resp = client.get("/export/?client_id=5")
        assert resp.status_code == 200
        mock_export.assert_called_once_with(mock_db, 5)
        mock_file.assert_called_once_with(mock_db, 5)

    def test_content_disposition_header(self, client):
        with patch(
            "app.routes.export.export_products_excel",
            return_value=self._mock_workbook(),
        ):
            with patch(
                "app.routes.export.get_master_filename",
                return_value="test_file.xlsx",
            ):
                resp = client.get("/export/")
        cd = resp.headers.get("content-disposition", "")
        assert "test_file.xlsx" in cd

    def test_response_is_valid_xlsx(self, client):
        """The response body should be a valid Excel file."""
        with patch(
            "app.routes.export.export_products_excel",
            return_value=self._mock_workbook(),
        ):
            with patch(
                "app.routes.export.get_master_filename",
                return_value="test.xlsx",
            ):
                resp = client.get("/export/")
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(resp.content))
        assert wb.active is not None
        assert wb.active.max_row >= 1

    def test_auth_required(self, unauth_client):
        """Export endpoint requires auth — should return 401."""
        with patch(
            "app.routes.export.export_products_excel",
            return_value=self._mock_workbook(),
        ):
            with patch(
                "app.routes.export.get_master_filename",
                return_value="test.xlsx",
            ):
                resp = unauth_client.get("/export/")
        assert resp.status_code == 401
