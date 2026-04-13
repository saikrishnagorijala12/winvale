import io
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import select, update, insert
from sqlalchemy.dialects.postgresql import insert as pg_insert
from fastapi import HTTPException
from app.models.client_profiles import ClientProfile
from app.models.client_contracts import ClientContracts
from app.models.product_master import ProductMaster
from app.models.product_history import ProductHistory
from app.models.product_dim import ProductDim
from app.utils.upload_helper import (
    MASTER_FIELDS,
    HISTORY_FIELDS,
    DIM_FIELDS,
    normalize,
    identity_signature,
    history_signature,
)
from app.redis_client import redis_client
from app.utils.cache import invalidate_pattern, invalidate_keys
from app.utils import s3_upload as s3

BATCH_SIZE = 500


def _is_effectively_blank_row(excel_row) -> bool:
    for cell in excel_row:
        if cell is None:
            continue
        if str(cell).strip() != "":
            return False
    return True


def upload_products(db: Session, client_id: int, file, user_email: str, progress_callback=None):

    if not db.query(ClientProfile.client_id).filter_by(client_id=client_id).first():
        raise HTTPException(status_code=404, detail="Client Not Found")

    if not db.query(ClientContracts.client_id).filter_by(client_id=client_id).first():
        raise HTTPException(status_code=404, detail="No Contract Found")

    file.file.seek(0)
    file_content = file.file.read()
    file.file.seek(0)
    
    wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    try:
        SHEET_NAME = "PRODUCTS"
        if SHEET_NAME not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Sheet '{SHEET_NAME}' not found. Please ensure your Excel file contains a sheet named '{SHEET_NAME}'.",
            )
        sheet = wb[SHEET_NAME]

        total_rows = sheet.max_row if sheet.max_row else 0
        
        rows = sheet.iter_rows(values_only=True)
        headers = None
        header_row_idx = 0
        for row in rows:
            header_row_idx += 1
            row_values = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "manufacturer" in row_values and "manufacturer_part_number" in row_values:
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                break

        if not headers:
            raise HTTPException(
                status_code=400,
                detail="Could not find header row. Please ensure 'manufacturer' and 'manufacturer_part_number' columns exist in your sheet.",
            )

        if total_rows > 0:
            total_rows = max(0, total_rows - header_row_idx)

        remaining_rows = sheet.iter_rows(
            min_row=header_row_idx + 1,
            values_only=True,
        )
        total_rows = sum(
            1 for excel_row in remaining_rows if not _is_effectively_blank_row(excel_row)
        )

        inserted = 0
        updated = 0
        reactivated = 0
        deleted = 0
        skipped = 0
        batch_counter = 0
        processed_count = 0

        
        existing_rows = db.execute(
            select(
                ProductMaster.manufacturer,
                ProductMaster.manufacturer_part_number,
                ProductMaster.product_id,
                ProductMaster.row_signature,
                ProductMaster.is_deleted,
                ProductDim.dim_id,
            )
            .outerjoin(ProductDim, ProductMaster.product_id == ProductDim.product_id)
            .where(ProductMaster.client_id == client_id)
        ).all()

        product_lookup = {
            (m, mpn): (pid, sig, is_del, did)
            for (m, mpn, pid, sig, is_del, did) in existing_rows
        }

        history_lookup = dict(
            db.execute(
                select(
                    ProductHistory.product_id,
                    ProductHistory.row_signature,
                ).where(
                    ProductHistory.client_id == client_id,
                    ProductHistory.is_current == True,
                )
            ).all()
        )

        seen_product_ids = set()

        current_headers = headers if headers else []
        batch_data = []
        for excel_row in rows:
            if _is_effectively_blank_row(excel_row):
                continue

            processed_count += 1
            # Initialize with defaults for all known fields to handle missing columns in Excel
            row_data = {f: normalize(None, f) for f in MASTER_FIELDS + DIM_FIELDS}
            
            for idx, col_name in enumerate(current_headers):
                if idx < len(excel_row):
                    row_data[col_name] = normalize(excel_row[idx], col_name)

            manufacturer = row_data.get("manufacturer")
            mpn = row_data.get("manufacturer_part_number")

            if not manufacturer or not mpn:
                skipped += 1
                continue

            batch_data.append(row_data)

            if len(batch_data) >= BATCH_SIZE:
                inserted, updated, reactivated, skipped = _process_batch(
                    db, client_id, batch_data, product_lookup, history_lookup, seen_product_ids,
                    inserted, updated, reactivated, skipped
                )
                db.commit()
                batch_data = []
                if progress_callback:
                    progress_callback(
                        processed_count=processed_count,
                        total_count=total_rows,
                        inserted=inserted,
                        updated=updated,
                        reactivated=reactivated,
                        skipped=skipped
                    )

        if batch_data:
            inserted, updated, reactivated, skipped = _process_batch(
                db, client_id, batch_data, product_lookup, history_lookup, seen_product_ids,
                inserted, updated, reactivated, skipped
            )
            db.commit()

        # Scenario 3: Soft-delete active products that were NOT present in the Excel.
        active_ids_before = {
            pid
            for (_, _, pid, _, is_del, _) in existing_rows
            if not is_del
        }
        to_delete_ids = active_ids_before - seen_product_ids

        if to_delete_ids:
            db.execute(
                update(ProductMaster)
                .where(ProductMaster.product_id.in_(to_delete_ids))
                .values(is_deleted=True)
            )
            deleted = len(to_delete_ids)
            db.commit()

        if progress_callback:
            progress_callback(
                processed_count=processed_count,
                total_count=total_rows,
                inserted=inserted,
                updated=updated,
                reactivated=reactivated,
                skipped=skipped
            )
    finally:
        wb.close()

    invalidate_pattern(redis_client, "products:all:*")
    invalidate_pattern(redis_client, f"products:client:{client_id}:*")
    invalidate_pattern(redis_client, "clients:all:*")
    invalidate_keys(redis_client, "clients:approved", f"clients:id:{client_id}")

    if inserted or updated or reactivated:
        file.file.seek(0)
        s3.save_uploaded_file(
            db, client_id, file, user_email, "gsa_upload"
        )

    return {
        "status_code": 201 if inserted or updated or reactivated else 200,
        "inserted": inserted,
        "updated": updated,
        "reactivated": reactivated,
        "deleted": deleted,
        "skipped": skipped,
    }


def _process_batch(
    db: Session,
    client_id: int,
    batch_data: list,
    product_lookup: dict,
    history_lookup: dict,
    seen_product_ids: set,
    inserted: int,
    updated: int,
    reactivated: int,
    skipped: int,
):
    counts = {
        "inserted": inserted,
        "updated": updated,
        "reactivated": reactivated,
        "skipped": skipped
    }

    to_insert_master = []
    to_update_master = []
    to_reactivate_master = []
    
    # Track row_data for insertions to create history/dim later
    insert_mappings = [] 
    
    update_history_ids = []
    new_history_mappings = []
    update_dim_mappings = []
    
    batch_new_keys = set() # To prevent double-insertion of same product in one batch

    for row_data in batch_data:
        manufacturer = str(row_data.get("manufacturer") or "")
        mpn = str(row_data.get("manufacturer_part_number") or "")
        identity_sig = identity_signature(row_data)
        history_sig = history_signature(row_data)

        existing = product_lookup.get((manufacturer, mpn))

        if not existing:
            if (manufacturer, mpn) in batch_new_keys:
                counts["skipped"] += 1
                continue
                
            batch_new_keys.add((manufacturer, mpn))
            
            master_payload = {f: row_data.get(f) for f in MASTER_FIELDS}
            master_payload.update({
                "client_id": client_id,
                "row_signature": identity_sig,
                "is_deleted": False
            })
            
            insert_mappings.append((row_data, master_payload, history_sig))
            to_insert_master.append(master_payload)
        else:
            product_id, _, is_del, did = existing
            seen_product_ids.add(product_id)

            if is_del:
                # Scenario 4: Reactivate
                master_payload = {f: row_data.get(f) for f in MASTER_FIELDS}
                master_payload.update({
                    "product_id": product_id,
                    "row_signature": identity_sig,
                    "is_deleted": False
                })
                to_reactivate_master.append(master_payload)
                
                update_history_ids.append(product_id)
                
                h_payload = {f: row_data.get(f) for f in HISTORY_FIELDS}
                h_payload.update({
                    "product_id": product_id,
                    "client_id": client_id,
                    "row_signature": history_sig,
                    "is_current": True,
                    "currency": str(row_data.get("currency") or "USD")
                })
                new_history_mappings.append(h_payload)

                d_payload = {f: row_data.get(f) for f in DIM_FIELDS}
                d_payload["product_id"] = product_id
                if did:
                    d_payload["dim_id"] = did
                update_dim_mappings.append(d_payload)

                history_lookup[product_id] = history_sig
                counts["reactivated"] += 1

            else:
                # Scenario 2: Update
                current_history_sig = history_lookup.get(product_id)
                if current_history_sig == history_sig:
                    counts["skipped"] += 1
                    continue

                update_history_ids.append(product_id)
                
                h_payload = {f: row_data.get(f) for f in HISTORY_FIELDS}
                h_payload.update({
                    "product_id": product_id,
                    "client_id": client_id,
                    "row_signature": history_sig,
                    "is_current": True,
                    "currency": str(row_data.get("currency") or "USD")
                })
                new_history_mappings.append(h_payload)

                m_payload = {f: row_data.get(f) for f in MASTER_FIELDS}
                m_payload.update({
                    "product_id": product_id,
                    "row_signature": identity_sig,
                    "is_deleted": False
                })
                to_update_master.append(m_payload)

                d_payload = {f: row_data.get(f) for f in DIM_FIELDS}
                d_payload.update({"product_id": product_id})
                if did:
                    d_payload.update({"dim_id": did})
                update_dim_mappings.append(d_payload)

                history_lookup[product_id] = history_sig
                counts["updated"] += 1

    # Execute Bulk Operations
    if to_insert_master:
        stmt = pg_insert(ProductMaster).values(to_insert_master).returning(
            ProductMaster.product_id, 
            ProductMaster.manufacturer, 
            ProductMaster.manufacturer_part_number
        )
        result = db.execute(stmt).all()
        
        id_map = {(str(m), str(mpn)): pid for pid, m, mpn in result}
        
        new_histories = []
        new_dims = []
        for row_data, _, h_sig in insert_mappings:
            m = str(row_data.get("manufacturer") or "")
            mpn = str(row_data.get("manufacturer_part_number") or "")
            pid = id_map.get((m, mpn))
            if pid:
                # Update lookups
                product_lookup[(m, mpn)] = (pid, identity_signature(row_data), False, None)
                history_lookup[pid] = h_sig
                seen_product_ids.add(pid)
                
                h_pay = {f: row_data.get(f) for f in HISTORY_FIELDS}
                h_pay.update({
                    "product_id": pid,
                    "client_id": client_id,
                    "row_signature": h_sig,
                    "is_current": True,
                    "currency": str(row_data.get("currency") or "USD")
                })
                new_histories.append(h_pay)

                d_pay = {f: row_data.get(f) for f in DIM_FIELDS}
                d_pay["product_id"] = pid
                new_dims.append(d_pay)
                counts["inserted"] += 1

        if new_histories:
            db.execute(insert(ProductHistory).values(new_histories))
        if new_dims:
            db.execute(insert(ProductDim).values(new_dims))

    if to_update_master:
        db.bulk_update_mappings(ProductMaster, to_update_master)
    
    if to_reactivate_master:
        db.bulk_update_mappings(ProductMaster, to_reactivate_master)

    if update_history_ids:
        db.execute(
            update(ProductHistory)
            .where(
                ProductHistory.product_id.in_(update_history_ids),
                ProductHistory.is_current == True
            )
            .values(is_current=False)
        )
    
    if new_history_mappings:
        db.execute(insert(ProductHistory).values(new_history_mappings))
    
    if update_dim_mappings:
        db.bulk_update_mappings(ProductDim, update_dim_mappings)

    return counts["inserted"], counts["updated"], counts["reactivated"], counts["skipped"]
