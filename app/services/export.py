from collections import defaultdict
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, select, text

from openpyxl import Workbook
from app.models.product_master import ProductMaster
from app.models.client_profiles import ClientProfile
from app.models.cpl_list import CPLList
from datetime import datetime, timezone
from typing import List, Optional
from app.models.modification_action import ModificationAction
from app.utils.export_constants import (
    PRICE_CHANGE_HEADERS, PRICE_CHANGE_GROUP_ROW, PRICE_CHANGE_MERGES,
    STANDARD_MODIFICATION_HEADERS, STANDARD_MODIFICATION_GROUP_ROW, STANDARD_MODIFICATION_MERGES,
    PRODUCT_EXPORT_HEADERS, PRODUCT_EXPORT_GROUP_ROW, PRODUCT_EXPORT_MERGES
)


def export_price_modifications_excel(
    db,
    client_id: Optional[int] = None,
    job_id: Optional[int] = None,
    selected_types: Optional[List[str]] = None,
    progress_callback=None,
):

    ordered_tabs = [
    ("NEW_PRODUCT", "Additions"),
    ("REMOVED_PRODUCT", "Deletions"),
    ("PRICE_INCREASE", "Price Increase"),
    ("PRICE_DECREASE", "Price Decrease"),
    ("DESCRIPTION_CHANGE", "Description Changes"),
    ("NO_CHANGE", "No Changes"),
    ]

    all_types = [t[0] for t in ordered_tabs]

    selected_types = [t for t in (selected_types or all_types) if t in all_types]

    if not selected_types:
        selected_types = all_types

    query = (
        db.query(ModificationAction)
        .options(
            joinedload(ModificationAction.product).joinedload(ProductMaster.dimension),
            joinedload(ModificationAction.cpl_item)
        )
        .filter(ModificationAction.action_type.in_(selected_types))
    )

    if client_id is not None:
        query = query.filter(ModificationAction.client_id == client_id)

    if job_id is not None:
        query = query.filter(ModificationAction.job_id == job_id)

    wb = Workbook(write_only=True)

    price_change_headers = PRICE_CHANGE_HEADERS
    price_change_group_row = PRICE_CHANGE_GROUP_ROW
    price_change_merges = PRICE_CHANGE_MERGES

    standard_headers = STANDARD_MODIFICATION_HEADERS
    standard_group_row = STANDARD_MODIFICATION_GROUP_ROW
    standard_merges = STANDARD_MODIFICATION_MERGES

    fill_dark = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_light = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")

    font_white = Font(name="Calibri", size=11, color="FFFFFF")
    font_standard = Font(name="Calibri", size=11)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    sheets = {}
    sheet_row_counters = {}

    def set_fixed_widths(ws, headers):
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def create_sheet(title, action_type):
        ws = wb.create_sheet(title)
        ws.freeze_panes = "D3"
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[2].height = 35

        is_price_tab = action_type in ["PRICE_INCREASE", "PRICE_DECREASE"]
        h_row_1 = price_change_group_row if is_price_tab else standard_group_row
        h_row_2 = price_change_headers if is_price_tab else standard_headers
        h_merges = price_change_merges if is_price_tab else standard_merges

        set_fixed_widths(ws, h_row_2)

        for start, end in h_merges:
            ws.merged_cells.add(f"{get_column_letter(start)}1:{get_column_letter(end)}1")

        # ---- Row 1 (Section Header) ----
        header_row_1 = []
        for value in h_row_1:
            cell = WriteOnlyCell(ws, value=value)
            cell.fill = fill_dark
            cell.font = font_white
            cell.alignment = center_align
            cell.border = header_border
            header_row_1.append(cell)
        ws.append(header_row_1)

        # ---- Row 2 (Column Headers) ----
        header_row_2 = []
        for value in h_row_2:
            cell = WriteOnlyCell(ws, value=value)
            cell.fill = fill_light
            cell.font = font_standard
            cell.alignment = center_align
            cell.border = header_border
            header_row_2.append(cell)
        ws.append(header_row_2)

        sheet_row_counters[action_type] = 3 
        return ws

    selected_types_set = set(selected_types)

    for action_type, title in ordered_tabs:
        if action_type in selected_types_set:
            sheets[action_type] = create_sheet(title, action_type)

    # Use Raw SQL for minimum overhead
    sql = text("""
        SELECT 
            m.*,
            p.item_type, p.manufacturer as p_mfr, p.manufacturer_part_number as p_mpn,
            p.vendor_part_number, p.sin, p.item_name as p_name, p.item_description as p_desc,
            p.recycled_content_percent, p.uom, p.quantity_per_pack, p.quantity_unit_uom,
            p.mfc_name, p.mfc_price, p.govt_price_no_fee, p.govt_price_with_fee,
            p.country_of_origin as p_coo, p.delivery_days, p.lead_time_code,
            p.fob_us, p.fob_ak, p.fob_hi, p.fob_pr, p.nsn, p.upc, p.unspsc,
            p.sale_price_with_fee, p.start_date, p.stop_date,
            p.default_photo, p.photo_2, p.photo_3, p.photo_4, p.product_url,
            p.warranty_period, p.warranty_unit_of_time,
            p.product_info_code, p.url_508, p.hazmat,
            p.dealer_cost, p.mfc_markup_percentage, p.govt_markup_percentage,
            d.length, d.width, d.height, d.physical_uom, d.weight_lbs,
            cpl.manufacturer_name as cpl_mfr, cpl.manufacturer_part_number as cpl_mpn,
            cpl.item_name as cpl_name, cpl.item_description as cpl_desc,
            cpl.origin_country as cpl_coo
        FROM modification_action m
        LEFT JOIN product_master p ON m.product_id = p.product_id
        LEFT JOIN product_dim d ON p.product_id = d.product_id
        LEFT JOIN cpl_list cpl ON m.cpl_id = cpl.cpl_id
        WHERE m.action_type IN :selected_types
        """ + (" AND m.client_id = :client_id" if client_id is not None else "") + """
        """ + (" AND m.job_id = :job_id" if job_id is not None else "") + """
    """)

    params = {
        "selected_types": tuple(selected_types),
        "client_id": client_id,
        "job_id": job_id
    }
    
    result = db.execute(sql, params)

    row_count = 0
    for res in result:
        row_count += 1
        if progress_callback and row_count % 1000 == 0:
            progress_callback(processed=row_count)

        ws = sheets.get(res.action_type)
        if not ws:
            continue

        is_price_tab = res.action_type in ["PRICE_INCREASE", "PRICE_DECREASE"]

        if is_price_tab:
            old_price = float(res.old_price) if res.old_price is not None else None
            new_price = float(res.new_price) if res.new_price is not None else None

            current_row = sheet_row_counters.get(res.action_type, 3)
            formula = f"=IFERROR(ABS((M{current_row}-L{current_row})/L{current_row}),0)"
            
            perc_change_cell = WriteOnlyCell(ws, value=formula)
            perc_change_cell.number_format = "0.00%"

            price_columns = [old_price, new_price, perc_change_cell]
        else:
            price = res.new_price if res.new_price is not None else res.old_price
            price_columns = [float(price) if price is not None else None]

        row = [
            res.item_type,
            res.p_mfr if res.p_mfr else res.cpl_mfr,
            res.p_mpn if res.p_mpn else res.cpl_mpn,
            res.vendor_part_number,
            res.sin,
            res.p_name if res.p_name else res.cpl_name,
            res.p_desc if res.p_desc else res.cpl_desc,
            float(res.recycled_content_percent) if res.recycled_content_percent is not None else None,
            res.uom,
            res.quantity_per_pack,
            res.quantity_unit_uom,
        ] + price_columns + [
            res.mfc_name,
            float(res.mfc_price) if res.mfc_price is not None else None,
            float(res.govt_price_no_fee) if res.govt_price_no_fee is not None else None,
            float(res.govt_price_with_fee) if res.govt_price_with_fee is not None else None,
            res.p_coo if res.p_coo else res.cpl_coo,
            res.delivery_days,
            res.lead_time_code,
            res.fob_us,
            res.fob_ak,
            res.fob_hi,
            res.fob_pr,
            res.nsn,
            res.upc,
            res.unspsc,
            float(res.sale_price_with_fee) if res.sale_price_with_fee is not None else None,
            res.start_date.isoformat() if res.start_date else None,
            res.stop_date.isoformat() if res.stop_date else None,
            res.default_photo,
            res.photo_2,
            res.photo_3,
            res.photo_4,
            res.product_url,
            res.warranty_period,
            res.warranty_unit_of_time,
            float(res.length) if res.length is not None else None,
            float(res.width) if res.width is not None else None,
            float(res.height) if res.height is not None else None,
            res.physical_uom,
            float(res.weight_lbs) if res.weight_lbs is not None else None,
            res.product_info_code,
            res.url_508,
            res.hazmat,
            float(res.dealer_cost) if res.dealer_cost is not None else None,
            float(res.mfc_markup_percentage) if res.mfc_markup_percentage is not None else None,
            float(res.govt_markup_percentage) if res.govt_markup_percentage is not None else None,
        ]

        ws.append(row)
        if res.action_type in sheet_row_counters:
            sheet_row_counters[res.action_type] += 1

    if progress_callback:
        progress_callback(processed=row_count)

    return wb



def export_products_excel(db, client_id: Optional[int] = None, progress_callback=None):

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("PRODUCTS")

    base_headers = PRODUCT_EXPORT_HEADERS[:]
    group_row = PRODUCT_EXPORT_GROUP_ROW[:]

    if client_id is None:
        base_headers.insert(0, "client_name")
        group_row.insert(0, "Client Details")

    merges = PRODUCT_EXPORT_MERGES

    # Shift merges if "Client Details" was inserted
    offset = 1 if client_id is None else 0
    adjusted_merges = [(start + offset, end + offset) for start, end in merges]

    total_columns = len(base_headers)

    ws.freeze_panes = get_column_letter(4 + offset) + "3"
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 35

    fill_dark = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_light = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")

    font_white = Font(name="Calibri", size=11, color="FFFFFF")
    font_standard = Font(name="Calibri", size=11)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col in range(1, total_columns + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    for start, end in adjusted_merges:
        ws.merged_cells.add(f"{get_column_letter(start)}1:{get_column_letter(end)}1")

    header_row_1 = []
    for value in group_row:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = fill_dark
        cell.font = font_white
        cell.alignment = center_align
        cell.border = header_border
        header_row_1.append(cell)

    ws.append(header_row_1)

    header_row_2 = []
    for value in base_headers:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = fill_light
        cell.font = font_standard
        cell.alignment = center_align
        cell.border = header_border
        header_row_2.append(cell)

    ws.append(header_row_2)

    # Use Raw SQL for maximum performance with 200k+ rows
    sql = text("""
        SELECT 
            p.*, 
            d.length as dim_length, d.width as dim_width, d.height as dim_height, 
            d.physical_uom as dim_uom, d.weight_lbs as dim_weight,
            c.company_name as client_name
        FROM product_master p
        LEFT JOIN product_dim d ON p.product_id = d.product_id
        LEFT JOIN client_profiles c ON p.client_id = c.client_id
        WHERE p.is_deleted = false
        """ + (" AND p.client_id = :client_id" if client_id is not None else "") + """
    """)
    
    params = {"client_id": client_id} if client_id is not None else {}
    result = db.execute(sql, params)

    row_count = 0
    for res in result:
        row_count += 1
        if progress_callback and row_count % 1000 == 0:
            progress_callback(processed=row_count)

        # Build row using raw result access
        prefix = [res.client_name] if client_id is None else []
        
        row = prefix + [
            res.item_type,
            res.manufacturer,
            res.manufacturer_part_number,
            res.vendor_part_number,
            res.sin,
            res.item_name,
            res.item_description,
            float(res.recycled_content_percent) if res.recycled_content_percent is not None else None,
            res.uom,
            res.quantity_per_pack,
            res.quantity_unit_uom,
            float(res.commercial_price) if res.commercial_price else None,
            res.mfc_name,
            float(res.mfc_price) if res.mfc_price else None,
            float(res.govt_price_no_fee) if res.govt_price_no_fee else None,
            float(res.govt_price_with_fee) if res.govt_price_with_fee else None,
            res.country_of_origin,
            res.delivery_days,
            res.lead_time_code,
            res.fob_us,
            res.fob_ak,
            res.fob_hi,
            res.fob_pr,
            res.nsn,
            res.upc,
            res.unspsc,
            float(res.sale_price_with_fee) if res.sale_price_with_fee else None,
            res.start_date.isoformat() if res.start_date else None,
            res.stop_date.isoformat() if res.stop_date else None,
            res.default_photo,
            res.photo_2,
            res.photo_3,
            res.photo_4,
            res.product_url,
            res.warranty_period,
            res.warranty_unit_of_time,
            float(res.dim_length) if res.dim_length is not None else None,
            float(res.dim_width) if res.dim_width is not None else None,
            float(res.dim_height) if res.dim_height is not None else None,
            res.dim_uom,
            float(res.dim_weight) if res.dim_weight is not None else None,
            res.product_info_code,
            res.url_508,
            res.hazmat,
            float(res.dealer_cost) if res.dealer_cost else None,
            float(res.mfc_markup_percentage) if res.mfc_markup_percentage else None,
            float(res.govt_markup_percentage) if res.govt_markup_percentage else None,
        ]

        ws.append(row)

    if progress_callback:
        progress_callback(processed=row_count)

    return wb



def get_master_filename(db: Session, client_id: Optional[int] = None):
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    if client_id is not None:
        client = db.query(ClientProfile).filter_by(client_id=client_id).first()
        if client:
            return f"{client.company_name}_products_{date_str}.xlsx"

    return f"all_products_{date_str}.xlsx"